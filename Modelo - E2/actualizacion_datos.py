from openpyxl import load_workbook, Workbook
from random import randint

import parametros as p
from paciente import Paciente
from creacion_datos import generacion_examenes, generacion_insumos, generación_disponibles, generacion_jornadas


def actualizar_datos(workbook, filename, hoja, lista_datos):
    pacientes = []
    for datos_paciente in lista_datos:
        pacientes.append(datos_paciente[0])

    for paciente in pacientes:
        celda = hoja.cell(row=1 + paciente, column=1)
        celda.value = paciente

        datos = lista_datos[paciente - 1]
        for dia in range(len(datos[1])):
            celda = hoja.cell(row=1 + paciente, column=1 + dia + 1)
            celda.value = datos[1][dia]

    workbook.save(filename)


def actualizar_disponibles(workbook, filename, lista_datos):
    pacientes = []
    for datos_paciente in lista_datos:
        pacientes.append(datos_paciente[0])

    for datos_paciente in lista_datos:
        n_paciente = datos_paciente[0]
        disponibles = datos_paciente[1]

        hoja_paciente = workbook.create_sheet(title=f"Paciente {n_paciente}")
        hoja_paciente['A1'] = "Médico"
        hoja_paciente['B1'] = "Día"
        hoja_paciente['C1'] = "Tiempo"

        for consulta in range(len(disponibles)):
            for item in range(3):
                celda = hoja_paciente.cell(
                    row=1 + consulta + 1, column=item + 1)
                celda.value = disponibles[consulta][item]

    workbook.save(filename)


def actualizar_medicos(workbook, filename, lista_datos):
    medicos = []
    for datos_medico in lista_datos:
        medicos.append(datos_medico[0])

    for datos_medico in lista_datos:
        n_medico = datos_medico[0]
        jornadas = datos_medico[1]

        hoja_medico = workbook.create_sheet(title=f"Medico {n_medico}")
        hoja_medico['A1'] = "Día"
        hoja_medico['B1'] = "Hora Entrada"
        hoja_medico['C1'] = "Hora Salida"

        for jornada in range(len(jornadas)):
            for item in range(3):
                celda = hoja_medico.cell(
                    row=1 + jornada + 1, column=item + 1)
                celda.value = jornadas[jornada][item]

    workbook.save(filename)


if __name__ == "__main__":

    # CREACIÓN DE DATOS SEMI-ALEATORIOS PARA EL MODELO

    # SETS
    P = range(1, p.Indices['n_P'] + 1)  # Pacientes
    M = range(1, p.Indices['M'] + 1)  # Médicos
    A = range(1, p.Indices['A'] + 1)  # PM (enfermero/a, anestesista, interno)
    D = range(1, p.Indices['D'] + 1)  # Días
    T = range(1, p.Indices['T'] + 1)  # Tiempo (en bloques de 15 min)
    # Tiempo prioritario
    T_PRI = range(p.Indices['T_PRI_min'], p.Indices['T_PRI_max'] + 1)

    # SETS ESPECIALES
    t_disponibles = range(p.Pacientes['disp_min'], p.Pacientes['disp_max'] + 1)
    t_examenes = range(p.MIN_EX, p.MAX_EX + 1)

    # ARCHIVOS
    #filename_pacientes = p.Paths['pacientes']
    #filename_medicos = p.Paths['medicos']
    filename_pacientes = "test_pacientes.xlsx"
    filename_medicos = "test_medicos.xlsx"

    # CREACIÓN DE EXCEL
    #EXCEL - PACIENTE
    datos_pacientes = Workbook()
    datos_examenes = datos_pacientes.create_sheet("Examenes")
    datos_examenes['A1'] = "Pacientes"

    datos_insumos = datos_pacientes.create_sheet("Insumos")
    datos_insumos['A1'] = "Pacientes"

    #EXCEL - PACIENTE
    datos_medicos = Workbook()

    # DATOS - PACIENTES
    # EXÁMENES
    examenes_validos = generacion_examenes(P, D, t_examenes)
    actualizar_datos(datos_pacientes, filename_pacientes,
                     datos_examenes, examenes_validos)

    # INSUMOS
    insumos_validos = generacion_insumos(P, D)
    actualizar_datos(datos_pacientes, filename_pacientes,
                     datos_insumos, insumos_validos)

    # DISPONIBLES
    disponibles_pacientes = generación_disponibles(
        P, M, D, t_disponibles)

    """for paciente in disponibles_pacientes:
        print(paciente[0])
        for horario in paciente[1]:
            print(horario)
        print()"""

    actualizar_disponibles(
        datos_pacientes, filename_pacientes, disponibles_pacientes)

    # DATOS - MEDICOS
    jornadas_medicos = generacion_jornadas(M, D)

    """for medico in jornadas_medicos:
        print(medico[0])
        for jornada in medico[1]:
            print(jornada)
        print()"""

    actualizar_medicos(datos_medicos, filename_medicos, jornadas_medicos)

    # Actualizar Datos (Guardar)
    datos_pacientes.save(filename_pacientes)
    datos_medicos.save(filename_medicos)
