import os
from openpyxl import load_workbook
from random import randint, choice, random

from paciente import Paciente


def carga_datos(filename, nombre_hoja, n_paciente):

    workbook = load_workbook(filename)

    hoja_paciente = workbook[nombre_hoja]
    max_row = hoja_paciente.max_row
    max_col = hoja_paciente.max_column

    for i in range(2, max_row + 1):
        celda = hoja_paciente.cell(row=i, column=1)
        if celda.value == n_paciente:

            lista_datos = []
            for j in range(2, max_col + 1):
                celda2 = hoja_paciente.cell(row=i, column=j)
                if celda2.value != None:
                    lista_datos.append(celda2.value)

            return lista_datos


def carga_disponibles(filename, n_paciente):

    workbook = load_workbook(filename)

    disponibles_paciente = workbook[f'Paciente {n_paciente}']
    max_row = disponibles_paciente.max_row
    columnas = range(1, 3 + 1)

    lista_disponibles = []

    for fila in range(2, max_row + 1):
        lista_valores = []

        for columna in columnas:
            celda = disponibles_paciente.cell(row=fila, column=columna)
            lista_valores.append(celda.value)

        lista_disponibles.append(lista_valores)

    return lista_disponibles


def carga_paciente(filename, paciente):

    paciente_examenes = carga_datos(filename, "Examenes", paciente)
    paciente_insumos = carga_datos(filename, "Insumos", paciente)
    paciente_disponibles = carga_disponibles(filename, paciente)

    paciente_prioritario = random() <= 0.2

    nuevo_paciente = Paciente(paciente, paciente_prioritario,
                              paciente_examenes, paciente_insumos, paciente_disponibles)

    return nuevo_paciente


def carga_medico(filename, medicos):

    workbook = load_workbook(filename)

    jornadas_medicos = []
    for medico in medicos:

        jornadas_medico = workbook[f'Medico {medico}']
        max_row = jornadas_medico.max_row
        columnas = range(1, 3 + 1)

        lista_jornadas = []

        for fila in range(2, max_row + 1):
            lista_valores = []

            for columna in columnas:
                celda = jornadas_medico.cell(row=fila, column=columna)
                lista_valores.append(celda.value)

            lista_jornadas.append(lista_valores)

        jornadas_medicos.append(lista_jornadas)

    return jornadas_medicos


def carga_jornadas(M, D, T, filename):
    JOR = {(m, d, t): 0 for m in M for d in D for t in T}

    lista_jornadas = carga_medico(filename, M)

    for m in M:
        # No puede iniciar su jornada despues de las 3 PM
        jornadas_medicos = lista_jornadas[m - 1]

        for jornada in jornadas_medicos:
            for t in range(jornada[1], jornada[2] + 1):
                JOR[m, jornadas_medicos.index(jornada) + 1, t] = 1

    return JOR


def carga_parametros(filename, index):
    workbook = load_workbook(filename)

    hoja = workbook['Parametros']
    return hoja[index].value


if __name__ == "__main__":
    pass
